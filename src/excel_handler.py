
import openpyxl
from pathlib import Path
from datetime import datetime

BASE_DIR = Path(__file__).parent.parent  # sube de src/ a ai-calls/
CLIENTES_PATH = BASE_DIR / "data" / "clientes.xlsx"
RESPUESTAS_PATH = BASE_DIR / "data" / "respuestas.xlsx"

def get_clientes() -> list[dict]:
    wb = openpyxl.load_workbook(CLIENTES_PATH)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    clientes = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):
            clientes.append(dict(zip(headers, row)))
    return clientes

def guardar_respuestas(cliente: dict, respuestas: list[dict]):
    if RESPUESTAS_PATH.exists():
        wb = openpyxl.load_workbook(RESPUESTAS_PATH)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Fecha", "Nombre", "Telefono", "Ubicacion", "Pregunta", "Respuesta"])

    for item in respuestas:
        ws.append([
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            cliente.get("nombre", ""),
            cliente.get("telefono", ""),
            cliente.get("ubicacion", ""),
            item.get("pregunta", ""),
            item.get("respuesta", ""),
        ])

    wb.save(RESPUESTAS_PATH)